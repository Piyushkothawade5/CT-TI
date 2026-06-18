import argparse
import json
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path

import openpyxl


CORE_FIELD_MAP = {
    "RATIO": "ratio",
    "Burden (VA)": "burden_va",
    "Accuracy Class": "accuracy_class",
    "ISF": "isf",
    "Min. Knee pt. volt.": "min_knee_pt_volt",
    "Max. Rct @ 75'c": "max_rct_75c",
    "Max. Exc. C/n. :- @VK/2": "max_exc_vk2",
    "Bare Core Dimensions": "bare_core_dim",
    "Core Material": "core_material",
    "Core weight (Kg)": "core_weight_kg",
    "Sec. Total Turns": "sec_total_turns",
    "Sec. Ter. Marking": "sec_ter_marking",
    "Sec. Conductor (S1-S2)": "sec_cond_s1s2",
    "Sec. Turns (S1-S2)": "sec_turns_s1s2",
    "Sec. Conductor (S2-S3)": "sec_cond_s2s3",
    "Sec. Turns (S2-S3)": "sec_turns_s2s3",
    "Sec. Conductor (S3-S4)": "sec_cond_s3s4",
    "Sec. Turns (S3-S4)": "sec_turns_s3s4",
    "Sec. Conductor (S4-S5)": "sec_cond_s4s5",
    "Sec. Turns (S4-S5)": "sec_turns_s4s5",
    "Sec. Copper weight (kg)": "sec_copper_wt",
    "Finished Core Dim. (mm)": "finished_core_dim",
    "Sec Connection": "sec_connection",
    "Wire Length": "wire_length",
    "Wire Colour": "wire_colour",
}

HEADER_MAP = {
    "Item No": "item_no",
    "CT Type": "ct_type",
    "Cust. Item No / Part code": "cust_part_code",
    "RATIO :-": "ratio",
    "RATED VOLTAGE": "rated_voltage",
    "STC": "stc",
    "I.L.": "insulation_level",
    "FREQ.": "frequency",
    "REF. STD.": "ref_std",
    "TI_No": "ti_no",
    "TI_Date": "ti_date",
    "Customer": "customer_name",
    "CUS_ORDER_DATE": "cus_order_date",
    "WO_No": "wo_number",
    "QTY": "quantity",
    "Sr_No": "serial_number",
    "CUS. ORDER. NO.": "cus_order_no",
    "PO ITEM NO.": "po_item_no",
}

LINE_SINGLE_MAP = {
    "CT final dim": "ct_final_dim",
    "GA Drg": "ga_drg",
    "INS CLASS": "ins_class",
    "PRI Turns": "pri_turns",
    "PRI Copper": "pri_copper",
    "Former": "former",
    "PRI Length": "pri_length",
    "PRI Weight": "pri_weight",
    "Sec. Terminal": "sec_terminal",
    "Ref TI:": "ref_ti",
    "Total Weight": "total_weight",
}

ITEM_COMMON_MAP = {
    "Item No": "item_no",
    "CT Type": "ct_type",
    "Cust. Item No / Part code": "cust_part_code",
    "RATIO :-": "ratio",
    "RATED VOLTAGE": "rated_voltage",
    "STC": "stc",
    "I.L.": "insulation_level",
    "FREQ.": "frequency",
    "REF. STD.": "ref_std",
    "CT final dim": "ct_final_dim",
    "GA Drg": "ga_drg",
    "INS CLASS": "ins_class",
    "PRI Turns": "pri_turns",
    "PRI Copper": "pri_copper",
    "Former": "former",
    "PRI Length": "pri_length",
    "PRI Weight": "pri_weight",
    "Sec. Terminal": "sec_terminal",
    "Customer": "default_customer",
}

DEFAULT_SIGNATURES = {
    "approved_by": "M.B",
    "checked_by": "L.A",
    "created_by": "R.P",
}

CUSTOMER_ALIASES = {
    "ALANAR": "ALFANAR",
    "ARABIAN GULF SWITCHGEAR": "ARABIAN GULF SWITCHGEAR",
    "BIN GHALIB": "BIN GHALIB",
    "CREATIVE ENGINNERS": "CREATIVE ENGINEERS",
    "LI=UCY NASHIK": "LUCY NASHIK",
    "LUCY DUBAI'": "LUCY DUBAI",
    "LUCY NASHIK'": "LUCY NASHIK",
    "LUCY NASHK": "LUCY NASHIK",
    "LUCY": "LUCY NASHIK",
    "MATELEC": "MATELEC",
    "MEGAWIN": "MEGAWIN",
    "POWER CONTROL ELECTRO SYSTEM": "POWER CONTROL ELECTRO SYSTEMS",
    "POWER CONTROL ELECTRO": "POWER CONTROL",
    "POWER CONTROL ELECTRO SYSTEMS": "POWER CONTROL",
    "STELMEC LIMITED": "STELMEC",
    "VOLTMAP TRANSFORMERS": "VOLTAMP TRANSFORMERS",
}

CT_TYPE_ALIASES = {
    "FG INSULATED CT": "FG TAPE INSULATED CT",
    "PLASIC CASE CT": "PLASTIC CASE CT",
    "PLASTIC CASE CT": "PLASTIC CASE CT",
    "PVC TAPE INSULATED CT.": "PVC TAPE INSULATED CT",
    "RESIN CASAT CT": "RESIN CAST CT",
    "RESIN CAST CT//////": "RESIN CAST CT",
    "RESIN CASTCT": "RESIN CAST CT",
    "RESIN INSULATED CT": "RESIN CAST CT",
    "TAPE INSULATED CT": "PVC TAPE INSULATED CT",
    "TAPE WOUND  CT": "PVC TAPE INSULATED CT",
    "TAPE WOUND CT": "PVC TAPE INSULATED CT",
    "TAPE WOUNDC CT": "PVC TAPE INSULATED CT",
}

INVALID_CT_TYPES = {
    "5010000386",
    "LUCY DUBAI",
}

ITEM_COLUMNS = [
    "item_no",
    "ct_type",
    "cust_part_code",
    "ratio",
    "rated_voltage",
    "stc",
    "insulation_level",
    "frequency",
    "ref_std",
    "core1",
    "core2",
    "core3",
    "ct_final_dim",
    "ga_drg",
    "ins_class",
    "ref_ti",
    "pri_turns",
    "pri_copper",
    "former",
    "pri_length",
    "pri_weight",
    "sec_terminal",
    "total_weight",
    "default_customer",
]

TI_COLUMNS = [
    "ti_no",
    "ti_date",
    "item_no",
    "wo_number",
    "customer_name",
    "cus_order_no",
    "cus_order_date",
    "quantity",
    "ct_type",
    "cust_part_code",
    "po_item_no",
    "serial_number",
    "ratio",
    "rated_voltage",
    "stc",
    "insulation_level",
    "frequency",
    "ref_std",
    "core1",
    "core2",
    "core3",
    "ct_final_dim",
    "ga_drg",
    "ins_class",
    "ref_ti",
    "pri_turns",
    "pri_copper",
    "former",
    "pri_length",
    "pri_weight",
    "sec_terminal",
    "total_weight",
    "created_by",
    "checked_by",
    "approved_by",
]


def clean_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return None if value == time(0, 0) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if value == 0:
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    if text in {"", "0", "00:00:00", "None"}:
        return None
    return text


def normalize_customer(value):
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("’", "'").strip()
    text = re.sub(r"\s+", " ", text).upper()
    return CUSTOMER_ALIASES.get(text, text)


def normalize_ct_type(value):
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"\s+", " ", text).upper().strip()
    text = CT_TYPE_ALIASES.get(text, text)
    if text in INVALID_CT_TYPES or text.isdigit():
        return None
    return text


def normalize_field(target, value):
    if target in {"customer_name", "default_customer"}:
        return normalize_customer(value)
    if target == "ct_type":
        return normalize_ct_type(value)
    return clean_text(value)


def clean_item_no(value):
    text = clean_text(value)
    if not text:
        return None
    return "".join(ch for ch in text if ch.isdigit())


def clean_date(value):
    text = clean_text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return text.split(" ")[0]


def first_present(current, new_value):
    return current if current not in (None, "") else new_value


def sql_string(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_json(value):
    return sql_string(json.dumps(value or {}, ensure_ascii=False, sort_keys=True)) + "::jsonb"


def row_dict(headers, row):
    return dict(zip(headers, row))


def rows_from_sheet(ws):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield row_dict(headers, row)


def core_index(core_name):
    text = clean_text(core_name)
    if text in {"Core 1", "1", "core1"}:
        return "core1"
    if text in {"Core 2", "2", "core2"}:
        return "core2"
    if text in {"Core 3", "3", "core3"}:
        return "core3"
    return None


def build_core(row):
    core = {}
    for source, target in CORE_FIELD_MAP.items():
        value = clean_text(row.get(source))
        if value is not None:
            core[target] = value
    if core.get("max_exc_vk2"):
        core["max_exc_is_vk2"] = "true"
    return core


def parse_items(wb):
    ws = wb["Item"]
    items = {}
    for row in rows_from_sheet(ws):
        item_no = clean_item_no(row.get("Item No"))
        if not item_no:
            continue
        item = items.setdefault(
            item_no,
            {"item_no": item_no, "core1": {}, "core2": {}, "core3": {}},
        )
        for source, target in ITEM_COMMON_MAP.items():
            value = clean_item_no(row.get(source)) if target == "item_no" else normalize_field(target, row.get(source))
            if value is not None:
                item[target] = first_present(item.get(target), value)
        ref_ti = clean_text(row.get("TI_No")) or clean_text(row.get("SourceSheet"))
        if ref_ti:
            item["ref_ti"] = first_present(item.get("ref_ti"), ref_ti)
        core_key = core_index(row.get("Core"))
        if core_key:
            item[core_key].update(build_core(row))
    return list(items.values())


def parse_transaction_records(wb):
    line_ws = wb["Transactions_Lines"]
    lines_by_header = defaultdict(list)
    for row in rows_from_sheet(line_ws):
        header_id = clean_text(row.get("HeaderID"))
        if header_id:
            lines_by_header[header_id].append(row)

    records_by_ti_no = {}
    header_ws = wb["Transactions_Header"]
    for row in rows_from_sheet(header_ws):
        ti_no = clean_text(row.get("TI_No"))
        if not ti_no:
            continue
        record = records_by_ti_no.setdefault(
            ti_no,
            {
                "ti_no": ti_no,
                "core1": {},
                "core2": {},
                "core3": {},
                **DEFAULT_SIGNATURES,
            },
        )
        for source, target in HEADER_MAP.items():
            if target == "item_no":
                value = clean_item_no(row.get(source))
            elif target in {"ti_date", "cus_order_date"}:
                value = clean_date(row.get(source))
            else:
                value = normalize_field(target, row.get(source))
            if value is not None:
                record[target] = first_present(record.get(target), value)

        for line in lines_by_header.get(clean_text(row.get("HeaderID")), []):
            core_key = core_index(line.get("Core"))
            if core_key:
                record[core_key].update(build_core(line))
            for source, target in LINE_SINGLE_MAP.items():
                value = clean_text(line.get(source))
                if value is not None:
                    record[target] = first_present(record.get(target), value)

    return list(records_by_ti_no.values())


def values_clause(row, columns):
    values = []
    for col in columns:
        if col in {"core1", "core2", "core3"}:
            values.append(sql_json(row.get(col)))
        elif col in {"ti_date", "cus_order_date"}:
            values.append(sql_string(row.get(col)))
        else:
            values.append(sql_string(row.get(col)))
    return "(" + ", ".join(values) + ")"


def write_insert(file, table, columns, rows, conflict_column):
    if not rows:
        return
    file.write(f"insert into public.{table} ({', '.join(columns)})\nvalues\n")
    file.write(",\n".join(values_clause(row, columns) for row in rows))
    updates = [
        f"{col} = excluded.{col}"
        for col in columns
        if col != conflict_column
    ]
    file.write(f"\non conflict ({conflict_column}) do update set\n  ")
    file.write(",\n  ".join(updates))
    file.write(";\n\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Path to Data.xlsx")
    parser.add_argument(
        "--output",
        default="supabase/import_data.sql",
        help="Output SQL file path",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    items = parse_items(wb)
    records = parse_transaction_records(wb)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as file:
        file.write("-- Generated from Data.xlsx. Run supabase/schema.sql before this file.\n")
        file.write("begin;\n\n")
        write_insert(file, "ct_items", ITEM_COLUMNS, items, "item_no")
        write_insert(file, "ct_ti_records", TI_COLUMNS, records, "ti_no")
        file.write(
            "update public.ct_ti_counter\n"
            "set current_value = coalesce((\n"
            "    select max((regexp_match(ti_no, '([0-9]+)$'))[1]::integer)\n"
            "    from public.ct_ti_records\n"
            "    where ti_no like left(public.format_ti_no(0), 11) || '%'\n"
            "      and ti_no ~ '[0-9]+$'\n"
            "  ), current_value);\n\n"
        )
        file.write("commit;\n")

    print(f"Wrote {output} with {len(items)} items and {len(records)} TI records.")


if __name__ == "__main__":
    main()
